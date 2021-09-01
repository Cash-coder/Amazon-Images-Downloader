from bs4 import BeautifulSoup as bs4
from requests_html import HTMLSession
import re

import xlsxwriter
from openpyxl.workbook.workbook import Workbook
from openpyxl import load_workbook


def make_query_url(item,attribute):
    query = item + ' ' + attribute
    #this is used for human reference in the file
    query_t = item + ' ' + attribute 
    query = query.replace(' ','+')

    url = 'https://www.amazon.es/s?k='+ query +'&__mk_es_ES=%C3%85M%C3%85%C5%BD%C3%95%C3%91&ref=nb_sb_noss'
    
    return url, query_t


def make_match_data(item,attribute):
    words = item.split(' ')
    new_words = []
    for word in words:
        word = word.capitalize() 
        new_words.append(word)
    item = ' '.join(new_words)

    if 'Iphone' in item:
        item = item.replace('Iphone','iPhone')

    attribute = attribute.capitalize()
    return item, attribute

def make_request(url):
    headers = {'user-agent':"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.159 Safari/537.36"}
    session = HTMLSession()
    r = session.get(url,headers=headers,allow_redirects=True)
    return r


def get_links(r):
    products_title = r.html.xpath('//div[@class="a-section a-spacing-none"]/div[@class="a-section a-spacing-none a-spacing-top-small"]/h2')
    #s = tag[0].text
    print()
    links = []
    for prod in products_title:
        #print(prod.text)
        if item_p and attribute_p in prod.text:
            link = prod.absolute_links
            print({'query':query,'link':link})
            entry = {'query':query,'link':link}
            links.append(entry)
    
    return links


def write_excel(links):
    wb = load_workbook(filename = 'pictures.xlsx')
    ws = wb.active

    for row in ws.iter_rows(values_only=True):
        if row[0] != None:
            pass




item = 'iPhone 12 pro' #Pro
attribute = 'grafito'

# _p means processed: from 'iphone pro' to 'iPhone Pro'
item_p,attribute_p = make_match_data(item,attribute)

url, query = make_query_url(item_p,attribute_p)

# r = make_request(url)
# status = r.status_code
# #To visualize response in the browser
# #save_html_response(r)
# if status == 200:
#     links = get_links(r)
# else:
#     print("Response status:{} for the query {} and url: {}".format(status,query,url))
    
links = ['lins1','lins2','lins3']

write_excel(links)