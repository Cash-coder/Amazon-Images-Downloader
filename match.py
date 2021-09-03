from twocaptcha import TwoCaptcha
from decimal import Decimal
from re import sub
from time import sleep

from bs4 import BeautifulSoup as bs4
from requests_html import HTMLSession
import re
import traceback

import xlsxwriter
from openpyxl.workbook.workbook import Workbook
from openpyxl import load_workbook


def make_query_url(item,attribute):
    query = item + ' ' + attribute
    #this is used for human reference in the file, with spaces instead of +
    query_t = item + ' ' + attribute
    query = query.replace(' ','+')

    #all deptartaments URL
    #url = 'https://www.amazon.es/s?k='+ query +'&__mk_es_ES=%C3%85M%C3%85%C5%BD%C3%95%C3%91&ref=nb_sb_noss'
    #Electronics URL
    url = 'https://www.amazon.es/s?k=' + query + '&i=electronics&__mk_es_ES=%C3%85M%C3%85%C5%BD%C3%95%C3%91&ref=nb_sb_noss_2'
    return url, query_t


#probably in disuse because now I lower() the prod_title
def make_match_data(item, attribute):

    if attribute:
        attribute_p = attribute.lower()
    if item:
        item_p = item.lower()

    #print(item_p,attribute_p)
    return item_p,attribute_p


def make_request(url):
    sleep(0.5) # used to avoid too much requests / second
    headers = {'user-agent':"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.159 Safari/537.36"}
    session = HTMLSession()

    proxies = {
        'http':'185.121.13.34',
        'http':'193.43.119.89',
        'http':'5.188.183.202',
        'http':'103.80.87.224',
        'http':'83.147.12.143',
        'http':'185.238.229.88',
        'http':'185.226.107.157',
        'http':'185.206.249.118',
        'http':'185.121.15.41',
        'http':'185.121.12.141',}

    r = session.get(url,headers=headers,proxies=proxies,allow_redirects=True)
    #print('request made: ',url)
    if r.status_code == 200:
        captcha = r.html.xpath('//h4[contains(text(),"Introduce los caracteres que se muestran a continuación")]')
        if captcha:
            print('---------detected CAPTCHA !!----------------')
            save_response(r) and print('saved response html')
            #solver = TwoCaptcha('c6aba1c6718b89d60b0d5f0c4eb34785')
            #result = solver.normal()
            #keep_page r.html.page to interact with the page
            r = session.get(url,headers=headers,proxies=proxies,allow_redirects=True)
        return r
    else:
        print('bad request', r.status_code)

nr = 0
def save_response(r):
    global nr
    with open('response'+str(nr)+'.html','wb') as f:
        f.write(r.content)
    nr += 1
    print('saved response')


def select(prod,title,query):
    links = []
    #print(query,title)
    if 'carcasa' not in title and 'funda' not in title and 'protector' not in title and 'soporte' not in title:
        link = prod.absolute_links
        link = str(link)
        link = link.replace('{','').replace('}','')
        link = link.replace("'",'')
        entry = [query,str(title),' ',str(link)]

        print('---------this entry was accepted:')
        print(entry)

        links.append(entry)

        return links

    else:
        print('not found from SELECT:','query:', query,'----','prod_title',title)
        print('-----------')
        write_no_results(query)


def get_matched_links(url,item_p,attribute_p,query,response):
    print('inside get matched links')
    #products_title = response.html.xpath('//div[@class="a-section a-spacing-none"]/div[@class="a-section a-spacing-none a-spacing-top-small"]/h2')
    products = response.html.xpath('//div[@data-component-type="s-search-result"]')

    # for p in products:
    #     title= p.xpath('//div[@class="a-section a-spacing-none a-spacing-top-small"]/h2')[0].text
    #     price = p.xpath('//span[@class="a-price-whole"]')
        # print(title)
        # print(price)
    #s = tag[0].text

    n_prods = (len(products))
    print('founded {} products'.format(n_prods))

    #prods
    '//div[@data-component-type="s-search-result"]'
    #price
    '//span[@class="a-price-whole"]'
    #save_response(response)
    links = []
    for prod in products:
        try:
            title= prod.xpath('//div[@class="a-section a-spacing-none a-spacing-top-small"]/h2')[0].text
            price = prod.xpath('//span[@class="a-price-whole"]')[0].text
            
            title = title.lower()
            price = price.split(',')[0]
            if '.' in price:
                n = price.replace('.',',')
                price = Decimal(sub(r'[^\d.]', '', n))
                price = int(price)
            else:
                price = Decimal(sub(r'[^\d.]', '', price))
                price = int(price)

            # try:
            #     price = int(price)
            # except:
            #     price = int(float(price))


            print('//price:',price,'  //title:', title)

            #no matter the order, if the words of the query are in title, include that url
            s = item_p.split(' ')
            n = len(s)
            n_t = len(attribute_p)

            #set the minimal price for the items, Example: No iphone costs less than 80€, but there unwanted are accesories
            min_price = 80
            if price < min_price:
                print('this -----PRICE----- is too low:',price)
                continue

            print('this is len:',n + n_t)
            if n == 1:
                if item_p in title and attribute_p in title :
                    link = select(prod,title,query)
                    links.append(link)

                else:
                    print('not found in get_MATCHED links:','n_prods:',n_prods,'---','query: ', query,'-----','prod_title:',title,'---', 'item: ',item_p,'---','attr: ',attribute_p,)
                    print('-----------')
                    write_no_results(query)

            elif n == 2:
                if attribute_p in title and s[0] in title and s[1] in title:
                    link = select(prod,title,query)
                    links.append(link)

                else:
                    print('not found in get_MATCHED links:','n_prods:',n_prods,'---','query: ', query,'-----','prod_title:',title,'---', 'item: ',item_p,'---','attr: ',attribute_p,)
                    print('-----------')
                    write_no_results(query)

            elif n == 3:
                if attribute_p in title and s[0] in title and s[1] in title and s[2] in title:
                    link = select(prod,title,query)
                    links.append(link)

                else:
                    print('not found in get_MATCHED links:','n_prods:',n_prods,'---','query: ', query,'-----','prod_title:',title,'---', 'item: ',item_p,'---','attr: ',attribute_p,)
                    print('-----------')
                    write_no_results(query)
            elif n == 4:
                if attribute_p in title and s[0] in title and s[1] in title and s[2] in title and s[3] in title:
                    link = select(prod,title,query)
                    links.append(link)

                else:
                    print('not found in get_MATCHED links:','n_prods:',n_prods,'---','query: ', query,'-----','prod_title:',title,'---', 'item: ',item_p,'---','attr: ',attribute_p,)
                    print('-----------')
                    write_no_results(query)
            elif n == 5:
                if attribute_p in title and s[0] in title and s[1] in title and s[2] in title and s[3] in title and s[4] in title:
                    link = select(prod,title,query)
                    links.append(link)

                else:
                    print('not found in get_MATCHED links:','n_prods:',n_prods,'---','query: ', query,'-----','prod_title:',title,'---', 'item: ',item_p,'---','attr: ',attribute_p,)
                    print('-----------')
                    write_no_results(query)

            elif n == 6:
                if attribute_p in title and s[0] in title and s[1] in title and s[2] in title and s[3] in title and s[4] in title  and s[5] in title:
                    link = select(prod,title,query)
                    links.append(link)
                else:
                    print('not found in get_MATCHED links:','n_prods:',n_prods,'---','query: ', query,'-----','prod_title:',title,'---', 'item: ',item_p,'---','attr: ',attribute_p,)
                    print('-----------')
                    write_no_results(query)
            elif n == 7:
                if attribute_p in title and s[0] in title and s[1] in title and s[2] in title and s[3] in title and s[4] in title  and s[5] in title  and s[6] in title in title:
                    link = select(prod,title,query)
                    links.append(link)
                else:
                    print('not found in get_MATCHED links:','n_prods:',n_prods,'---','query: ', query,'-----','prod_title:',title,'---', 'item: ',item_p,'---','attr: ',attribute_p,)
                    print('-----------')
                    write_no_results(query)
            elif n == 8:
                if attribute_p in title and s[0] in title and s[1] in title and s[2] in title and s[3] in title and s[4] in title  and s[5] in title  and s[6] in title and s[7] in title:
                    link = select(prod,title,query)
                    links.append(link)
                else:
                    print('not found in get_MATCHED links:','n_prods:',n_prods,'---','query: ', query,'-----','prod_title:',title,'---', 'item: ',item_p,'---','attr: ',attribute_p,)
                    print('-----------')
                    write_no_results(query)

            else:
                print('not found in get_MATCHED links:','n_prods:',n_prods,'---','query: ', query,'-----','prod_title:',title,'---', 'item: ',item_p,'---','attr: ',attribute_p,)
                print('-----------')
                write_no_results(query)

            return(links)

        except Exception as e:
            print('error in assign title and price, this is the url:')
            print(url)
            print(e)
            print('-------This is the prod data with the problem:-----------')
            print(prod.text)
            continue


        # if item_p in title and attribute_p in title:
        #     #if prod.text not in ['Carcasa', 'Funda', 'Protector', 'Soporte'] :
        #     #if 'Carcasa' and 'Funda' and 'Protector' and 'Soporte' not in prod.text:
        #     if 'carcasa' not in title and 'funda' not in title and 'protector' not in title and 'soporte' not in title:
        #         link = prod.absolute_links
        #         link = str(link)
        #         link = link.replace('{','').replace('}','')
        #         link = link.replace("'",'')
        #         #print({'query':query,'link':link})
        #         #entry = {'query':query,'link':link,'prod_title':prod.text}
        #         entry = (query,prod.text,' ',link)
        #         print(entry)
        #         links.append(entry)




def write_excel(links):
    try:
        wb = load_workbook(filename = 'matches.xlsx')
        #wb = Workbook()
        ws = wb.active

        for entry in links:
            ws.append(entry)
            # #print(query,link,prod_title)
        separator = ('################################################','################################################','################################################','################################################')
        ws.append(separator)

        wb.save('matches.xlsx')
    except Exception as e:
        print(e)
        pass

def write_no_results(query):

    wb = load_workbook(filename = 'no_results.xlsx')
    #wb = Workbook()
    ws = wb.active
    entry = (query,'something_here')
    ws.append(entry)
    wb.save('no_results.xlsx')


def get_item_attribute():
    item_attribute_list = []
    wb = load_workbook(filename = 'phones_color_variations.xlsx')
    ws = wb.active

    for row in ws.iter_rows(values_only=True):
        item = row[0]
        attribute = row[1]
        item_attribute_list.append({'item':item,'attribute':attribute})
    return item_attribute_list


item_attribute_list = get_item_attribute()
for element in item_attribute_list:

    item = element.get('item')
    attribute = element.get('attribute')

    #print('1',item,attribute)

    #process the item and the attribute to match Amazon's standart (Capitalization, iPhone,etc...)
    #This is used later to identify matches within the titles of the prods.
    # _p means processed: from 'iphone pro' to 'iPhone Pro'
    item_p,attribute_p = make_match_data(item,attribute)
    #print('2',item_p,attribute_p)

    #with the above data make the url and the query (used later in the excel)
    url, query = make_query_url(item_p,attribute_p)

    #make the request with the query
    try:
        response = make_request(url)
    except:
        continue

    #extract the links of the products which titles matches the query
    #list of dicts with link , query, prod_title
    links = get_matched_links(item_p=item_p,attribute_p=attribute_p,query=query,url=url,response=response)
    # write excel with query , prod_title , selection, link
    #selection is if the human validate that url has the needed pictures
    write_excel(links)
