from typing import Match
from openpyxl.workbook.workbook import Workbook
from openpyxl import load_workbook

# list_a = ['iphone 12 pro max oro','iphone 12 pro max plata' ]
# list_b = ['iphone 12 pro max oro' ]

# for item in list_a:
#     if item not in list_b:
#         print(item)

''' given 2 files on complete and other with missing items,
 it creates a new file with the missing items
 
 when invoek the functions below you can specify the column pf the data'''

matches = 'phones_matches.xlsx'
complete = 'phones_color_variations.xlsx'
missing_items = 'phones_missing.xlsx'

#given an excel file it returns a list of the values
def read(filename, column=0):
    wb = load_workbook(filename = filename)
    ws = wb.active

    complete_list = []
    for row in ws.iter_rows(values_only=True):
        item = row[column]
        try:
            item = item.lower()
        except:
            pass
        complete_list.append(item)
    
    return complete_list

#given 2 lists it writes the missing elements to a third file
def write_missing(complete_list, incomplete_list):
    wb = Workbook()
    ws = wb.active
    n = 1
    for item in complete_list:
        if item not in incomplete_list:
            print('This item is missing in INcomplete_list:  ', item)
            ws.cell(row=n, column=1, value=item)
            n += 1
    wb.save(missing_items)
    

complete_list = read(complete,column=2)
matches_list = read(matches)

write_missing(complete_list, matches_list)






############

# wb = load_workbook(filename = complete)
# ws = wb.active

# complete_list = []

# for row in ws.iter_rows(values_only=True):
#     item = row[1]
#     complete_list.append(item)