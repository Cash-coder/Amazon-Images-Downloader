# removes unwanted elements from variation title, like '()' to make the match eaasier
# '"' for 'pulgadas'
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

input_file = 'mac_source.xlsx'
output_file = 'mac_p.xlsx'

#outputs list of cleaned items
def clean():
    wb = load_workbook(filename = input_file)
    ws = wb.active
    cleaned = []
    #item_attribute_list = []
    for row in ws.iter_rows(values_only=True):
        item = row[0]
        item = item.replace('(','').replace(')','') #ipad pro (2021)
        item = item.replace('"',' pulgadas') # " for ' pulgadas'
        cleaned.append(item)
        # if 'GB' in item:
        #     chain = item.split(' ')
        #     item_p = chain.remove(chain[-1])
        #     item_p = chain.remove(chain[-2])
        #     print(item_p)
        #     #myList.remove(myList[len(myList)-1])
    return cleaned

def write(cleaned):
    wb = Workbook()
    ws = wb.active

    n = 1
    for item in cleaned:
        ws.cell(row=n, column=1, value=item)
        n += 1
    
    wb.save(output_file)


cleaned = clean()
print(cleaned)
write(cleaned)