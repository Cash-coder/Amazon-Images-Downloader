from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

item_attribute_list = []
wb = load_workbook(filename = 'tablets.xlsx')
ws = wb.active

for row in ws.iter_rows(values_only=True):
    item = row[0]
    item_p = item_p.replace('(','').replace(')','') #ipad pro (2021)

    if 'GB' in item:
        chain = item.split(' ')
        item_p = chain.remove(chain[-1])
        item_p = chain.remove(chain[-2])

        print(item_p)
        #myList.remove(myList[len(myList)-1])