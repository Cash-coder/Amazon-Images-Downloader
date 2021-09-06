import undetected_chromedriver.v2 as uc
from selenium.webdriver.common.by import By

from decimal import Decimal
from re import sub
from time import sleep
from openpyxl.workbook.workbook import Workbook
from openpyxl import load_workbook

input_file = 'hp_lap.xlsx' #source file to make the requests
output_file = 'hp_matches.xlsx' #used to save the results
no_results = 'hp_noresults.xlsx' 

filter_price = 120 #used to filter matches also by price, no iphone X costs less than 120€, filter other accesories

#get item + color from xlsx list
def get_target_list():
    item_attribute_list = []
    wb = load_workbook(filename = input_file) 
    ws = wb.active

    for row in ws.iter_rows(values_only=True):
        item = row[0]
        attribute = row[1]
        if attribute == None: # phones-tablets with only one color, use a white space to keep the code as it is
            attribute = ' '
            print('###################This item hasn\'t attribute!!',item)
        item_attribute_list.append({'item':item,'attribute':attribute})
    return item_attribute_list

#lower and join items and colors to compare
def make_match_data(item, attribute):
    
    if item:
        item_p = str(item)
        item_p = item_p.lower()
        item_p = item_p.replace('(','').replace(')','') #already done in clenaer.py

    #sometimes the attribute is an integer, like 2019, can't apply lower()
    if attribute:
        try:
            attribute_p = str(attribute)    
            attribute_p = attribute_p.lower()
        except Exception as e:
            attribute_p = str(attribute)
            print(e)
            pass
    #print(item_p,attribute_p)
    return item_p, attribute_p

#given the item and the color, make the url to get
def make_query_url(item,attribute):

    if attribute: #sometimes there's no attribute, like "hp 15s" 
        query = item + ' ' + attribute
        #query_t is used later for human reference in the file, with spaces instead of +, to know which query it is, 'iphone 12 verde'
        query_t = item + ' ' + attribute
        query = query.replace(' ','+')
        
    else:
        query = item
        query = query.replace(' ','+')

    url = 'https://www.amazon.es/s?k=' + query + '&i=electronics&__mk_es_ES=%C3%85M%C3%85%C5%BD%C3%95%C3%91&ref=nb_sb_noss_2'
    #in Amazon all depatartaments URL
    #url = 'https://www.amazon.es/s?k='+ query +'&__mk_es_ES=%C3%85M%C3%85%C5%BD%C3%95%C3%91&ref=nb_sb_noss'
    #Electronics URL
    
    return url, query_t

def make_request(url):
    options = uc.ChromeOptions() 
    options.add_argument('--headless') 
    options.add_argument('--disable-gpu') 
    d = uc.Chrome(options=options)
    with d:
        d.get(url)
        print('Request sent, query: ', query)
        #if the browser opens 2 tabs
        if len(d.window_handles) > 1:
            d.switch_to.window(d.window_handles[1])
        sleep(3)
        prods = d.find_elements_by_xpath('//div[@data-component-type="s-search-result"]')
        print('founded {} prods'.format(len(prods)))
        return prods


def get_prod_matches(prods,item_p,attribute_p,query):
    data_list = []
    for p in prods:
        try:
            title = p.find_element_by_xpath('.//div[@class="a-section a-spacing-none a-spacing-top-small"]/h2').text
            title = title.lower()
            try:
                raw_price = p.find_element_by_xpath('.//span[@class="a-price-whole"]').text
                price = fix_price(raw_price)
                #print(title)
            except:
                price = 10_000
            #set the minimal price for the items, Example: No iphone costs less than 80€, but there unwanted are accesories
            min_price = filter_price
            if price < min_price:
                #print('this -----PRICE----- is too low:',price, title)
                continue
            
            
            #split the item into words, to search standalone words in the titles instead only one string
            s = item_p.split(' ')
            n = len(s)
            n_t = len(attribute_p)

            if n == 1:
                if item_p in title and attribute_p in title :
                    entry = select(p,title,query)
                    data_list.append(entry) # list of dicts that contains accepted query, title, link

                else:
                    print('not found in get_MATCHED entrys:','---','query: ', query,'-----','prod_title:',title,'---', 'item: ',item_p,'---','attr: ',attribute_p,)
                    print('-----------')
                    ##write_no_results(query)

            elif n == 2:
                if attribute_p in title and s[0] in title and s[1] in title:
                    entry = select(p,title,query)
                    data_list.append(entry) # list of dicts that contains accepted query, title, link

                else:
                    print('not found in get_MATCHED data:','---','query: ', query,'-----','prod_title:',title,'---', 'item: ',item_p,'---','attr: ',attribute_p,)
                    print('-----------')
                    #write_no_results(query)

            elif n == 3:
                if attribute_p in title and s[0] in title and s[1] in title and s[2] in title:
                    entry = select(p,title,query)
                    data_list.append(entry) # list of dicts that contains accepted query, title, link

                else:
                    print('not found in get_MATCHED data:','---','query: ', query,'-----','prod_title:',title,'---', 'item: ',item_p,'---','attr: ',attribute_p,)
                    print('-----------')
                    #write_no_results(query)
            elif n == 4:
                if attribute_p in title and s[0] in title and s[1] in title and s[2] in title and s[3] in title:
                    entry = select(p,title,query)
                    data_list.append(entry) # list of dicts that contains accepted query, title, link

                else:
                    print('not found in get_MATCHED data:','---','query: ', query,'-----','prod_title:',title,'---', 'item: ',item_p,'---','attr: ',attribute_p,)
                    print('-----------')
                    #write_no_results(query)
            elif n == 5:
                if attribute_p in title and s[0] in title and s[1] in title and s[2] in title and s[3] in title and s[4] in title:
                    entry = select(p,title,query)
                    data_list.append(entry) # list of dicts that contains accepted query, title, link

                else:
                    print('not found in get_MATCHED data:','---','query: ', query,'-----','prod_title:',title,'---', 'item: ',item_p,'---','attr: ',attribute_p,)
                    print('-----------')
                    #write_no_results(query)

            elif n == 6:
                if attribute_p in title and s[0] in title and s[1] in title and s[2] in title and s[3] in title and s[4] in title  and s[5] in title:
                    entry = select(p,title,query)
                    data_list.append(entry) # list of dicts that contains accepted query, title, link
                else:
                    print('not found in get_MATCHED data:','---','query: ', query,'-----','prod_title:',title,'---', 'item: ',item_p,'---','attr: ',attribute_p,)
                    print('-----------')
                    #write_no_results(query)
            elif n == 7:
                if attribute_p in title and s[0] in title and s[1] in title and s[2] in title and s[3] in title and s[4] in title  and s[5] in title  and s[6] in title in title:
                    entry = select(p,title,query)
                    data_list.append(entry) # list of dicts that contains accepted query, title, link
                else:
                    print('not found in get_MATCHED data:','---','query: ', query,'-----','prod_title:',title,'---', 'item: ',item_p,'---','attr: ',attribute_p,)
                    print('-----------')
                    #write_no_results(query)
            elif n == 8:
                if attribute_p in title and s[0] in title and s[1] in title and s[2] in title and s[3] in title and s[4] in title  and s[5] in title  and s[6] in title and s[7] in title:
                    entry = select(p,title,query)
                    data_list.append(entry)
                    #return(data_list) # list of dicts that contains accepted query, title, link
                else:
                    print('not found in get_MATCHED data:','---','query: ', query,'-----','prod_title:',title,'---', 'item: ',item_p,'---','attr: ',attribute_p,)
                    print('-----------')
                    ##write_no_results(query)
        except Exception as e: 
            print('Exception in get_prod_matches(), this product')
            #print('query:',query,'---------','title:',title)
            print(e)
            continue

    return(data_list)


#example: from 1.250,00 to 1250
def fix_price(price):
    #print('input price in fix_price:',price)
    price = price.split(',')[0]
    if '.' in price:
        n = price.replace('.',',')
        price = Decimal(sub(r'[^\d.]', '', n))
        price = int(price)
    else:
        price = Decimal(sub(r'[^\d.]', '', price))
        price = int(price)
    #print('output price in fix_price',price)
    return price


def select(prod,title,query):
    #print(query,title)
    if 'carcasa' not in title and 'funda' not in title and 'protector' not in title and 'soporte' not in title:
        #prod.find_element(By.XPATH, '//button[text()="Some text"]')
        #link = prod.absolute_links
        link = prod.find_element_by_tag_name('a').get_attribute('href')
        link = str(link)
        entry = {'query':query,'title':title,'link':link}
        print('---------this entry was accepted:')
        print(entry)
        return entry


row = 1
row_2 = 1
def write_excel(data_list,query):
    global row
    global row_2

    if data_list:
        print('XXXXXXXXXXXXXXXXXXXXXXX') #if there are results for this query, write to matches.xlsx
        print('there is a data list') #if there are results for this query, write to matches.xlsx
        print(data_list)
        for data in data_list:
            try:
                #wb = Workbook()
                wb = load_workbook(filename = output_file)
                ws = wb.active
                
                query = data.get('query')
                title = data.get('title')
                link = data.get('link')

                #query
                ws.cell(row=row,column=1,value=query)
                #title
                ws.cell(row=row,column=2,value=title)
                #white space
                ws.cell(row=row,column=3,value=' ')
                #link
                ws.cell(row=row,column=4,value=link)
                
                print('written cell in row',row)
                row += 1
                #ws.append(entry)
                # #print(query,link,prod_title)
                wb.save(output_file)
            except Exception as e:
                print(e)
                pass
            
        print('going to write separator in line',row)
        separator = ('############################################################################')
        wb = load_workbook(filename = output_file)
        ws = wb.active
        ws.cell(row=row,column=1,value= '############################################################################' )
        wb.save('tablets_matches.xlsx')
        row += 1
        print('saved file')
    
    else: #if there aren't results (empty list) for this query, write to no_results file
        print('AAAAAAAAAAAAAAA') #if there are results for this query,
        print('there ARE NOT a data list') #if there are results
        print(data_list)
        #wb = Workbook()
        write_no_results(query) 
        #wb = load_workbook(filename = 'tablets_no_results.xlsx')
        #ws = wb.active
        #ws.cell(row=row_2,column=1,value=query)
        #wb.save('no_results.xlsx')


############
row_2 = 1
def write_no_results(query):
    global row_2
    #wb = Workbook()
    wb = load_workbook(filename = no_results)
    ws = wb.active
    ws.cell(row=row_2, column=1, value=query)
    row_2 += 1
    #entry = (query,'something_here')
    #ws.append(entry)
    wb.save(no_results)


target_list = get_target_list()
for e in target_list: #[50:60]
    try:
        #get the data from the list
        item = e.get('item')
        color = e.get('attribute')
        #lower() character to compare later with prod_title
        item_p,attribute_p = make_match_data(item,color)
        #make the url to get, and get the query text for reference
        url, query = make_query_url(item_p,attribute_p)
        #make the get with selenium
        try:
            prods = make_request(url)
        except:
            continue
        
        matches = get_prod_matches(prods=prods, item_p=item_p, attribute_p=attribute_p, query=query)
        write_excel(matches,query)
    except Exception as e:
        print(e)
        continue