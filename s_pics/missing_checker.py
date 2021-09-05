from openpyxl.workbook.workbook import Workbook
from openpyxl import load_workbook

#given an excel file it returns a list of the values
def read(filename):
    wb = load_workbook(filename = filename)
    ws = wb.active

    complete_list = []
    for row in ws.iter_rows(values_only=True):
        item = row[0]
        try:
            item = item.lower()
        except:
            pass
        complete_list.append(item)
    
    return complete_list

#given 2 lists it writes the missing elements to a third file
def write_missing(complete_list, incomplete_list, new_list_name):
    wb = Workbook()
    ws = wb.active

    n = 1
    for item in complete_list:
        if item not in matches_list:
            ws.cell(row=n, column=1, value=item)
        n += 1
    
    wb.save(new_list_name)
    
#save file here

matches = 'tablets_matches.xlsx'
complete = 'tablets_color_variations.xlsx'

complete_list = read(complete)
matches_list = read(matches)

write_missing(complete_list, matches_list, 'missing_tablets.xlsx')






############

# wb = load_workbook(filename = complete)
# ws = wb.active

# complete_list = []

# for row in ws.iter_rows(values_only=True):
#     item = row[1]
#     complete_list.append(item)