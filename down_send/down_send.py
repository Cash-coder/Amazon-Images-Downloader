#https://pypi.org/project/mega.py/
from csv import excel
from mega import Mega
import login_file # .py file with MEGA user and password
from requests_html import HTMLSession
import requests
import re
from time import sleep
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
import os
import traceback
import json
import shutil
import logging


#logging.basicConfig(level=logging.INFO)
# logging.info('got mega credentials')
# debug
# info
# warning
# error
# critical

#setting pics file path
def set_path(machine):
    if machine == 'windows' : pics_folder = r'C:/Users/HP EliteBook/OneDrive/A_Miscalaneus/Escritorio/Code/git_folder/images_downloader/down_send/pics_folder/'
    elif machine == 'server': pics_folder = '/home/nonroot/pics/down_send/pics_folder/'
    return pics_folder


#init mega
email = login_file.email
password = login_file.passw
logging.info('got mega credentials')

machine = 'windows' # 'server' or 'windows'
input_file  = 'set_laptops_mac.xlsx'
output_file = 'processed_laptops_mac.xlsx'
errors_file = 'errors_laptops_mac.xlsx'

pics_folder = set_path(machine)
errors_folder = 'errors_mega_uploads'



#returns a list of dicts with {item , url}
def extract():
    wb = load_workbook(filename = input_file)
    ws = wb.active
    target_list = []
    #item_attribute_list = []
    for row in ws.iter_rows(values_only=True):
        item = row[0]
        url = row[1]

        entry = {'item':item, 'url':url}
        target_list.append(entry)
        # if 'GB' in item:
        #     chain = item.split(' ')
        #     item_p = chain.remove(chain[-1])
        #     item_p = chain.remove(chain[-2])
        #     print(item_p)
        #     #myList.remove(myList[len(myList)-1])
    logging.info('extracted excel')
    return target_list


n_errors = 1
def write_bad_result(url,item='no item appended'):
    ''' URL ITEM'''
    global n_errors

    wb = load_workbook(filename = errors_file)
    ws = wb.active

    ws.cell(row=n_errors,column=1,value=item)
    ws.cell(row=n_errors,column=2,value=url)
    wb.save(errors_file)

    n_errors += 1
    print('written bad result', item)


def download_picture(url, item, n): # Download To local machine
    global pics_folder
    try:
        pic = proxy_request(url, item) 
        
        if pic.status_code == 200:
            file_name = item + ' ' + str(n) +'.jpg'
            file_name = file_name.replace(' ','_')
            if '__' in file_name: file_name = file_name.replace('__','_')
            if '___' in file_name: file_name = file_name.replace('___','_')
            #some items has / in the name: like iphone 12 negro/grafito
            if '/' in file_name : file_name = file_name.replace('/','&')
    
            try:
                complete_path = pics_folder + file_name
                with open(complete_path, 'wb') as f:
                    f.write(pic.content)
                    logging.debug(f'downloaded this pic: {file_name}')
                    #print('downloaded this pic: ',file_name)
            except Exception as e:
                print(e)
                print('failed to save this pic',item, url)
                write_bad_result(item, url)
        else:
            print('status code for this picture (not item) is {} for this url: {}'.format(pic.status_code, url))
            write_bad_result(item, url)
            
    except Exception as e:
        print('exception downloading a file :', e, 'satuscode:',pic.status_code)
        #save_html_response(pic,item)
        write_bad_result(item,url)


        
def save_html_response(r, item='-'):
    item = item.replace(' ', '_')
    with open(item + 'response.html','w',encoding='utf-8') as f:
        f.write(r.text)


def get_pictures_urls(url, r):
    ''' returns a list with image URL's picks the url from Amazon or Backmarket'''
    try:
        url_list = []
        if r == 'request error':
            return 'request error'

        # Backmarket image links picker
        if 'backmarket' in url:
            tags = r.html.xpath('//div[@class="flex justify-center md:hidden"]/div/div/ul/li/img')
            search_pattern = r'https(.*?).jpg'
            for tag in tags:
                # print(tag)
                url = re.search(search_pattern, str(tag))
                # print(url)
                if url != None:
                    url = url.group()
                    url_list.append(url)
                else: 
                    print('no match in this regex, url: ', url)
                    continue


            #used to log the process
            total_pics = len(r.html.xpath('//div[@class="hidden md:block"]//descendant::ul/li')) #measured by the number of thumbnail pictures in prod page
            total_pics = total_pics / 2 # trick to get exact number
            total_pics = int(total_pics) # trick to get exact number
            #print(f'item: {item}; total pics: {total_pics}; ')

        # Amazon image links picker
        elif 'amazon' in url:
            tag = r.html.find('script', containing='P.when(\'A\').register("ImageBlockATF", function(A){')
            s = tag[0].text

            total_pics = r.html.xpath('//li[@class="a-spacing-small item imageThumbnail a-declarative"]') #measured by the number of thumbnail pictures in prod page
            total_pics = len(r.html.xpath('//span[@class="a-button a-button-thumbnail a-button-toggle"]/span[@class="a-button-inner"]/input')) #measured by the number of thumbnail pictures in prod page
            total_pics = total_pics - 1 #In amazon the html tag displays one more
            logging.debug('Total pics finded in Amazon:  {total_pics}')
            #total_pics = r.html.xpath('//li[@class="a-spacing-small item imageThumbnail a-declarative"]') #measured by the number of thumbnail pictures in prod page
            #print(f'item: {item}; total pics: {total_pics}; ')

            # hiRes == High Resolution 
            # main  ==  ?
            # large ==  ?
            matches = re.findall(r'hiRes(.*?).jpg', s)
            for match in matches:
                #print(match,'\n')
                if match == None :
                    print(' NONE in this match: ', match)
                    continue
                #some pics haven't HiRes pics, use large ones
                elif ':null,"thumb":' in match : #Amazon hasn't hires pic, chose large ones
                    #print('founded NULL match, this :', match)
                    continue
                    # alternative_matches = re.findall(r'large(.*?).jpg', s)
                    # for match in alternative_matches:
                        # #print(f'this is complete match: {match}')
                        # if '{' in match: url = match.replace('":{"','') + '.jpg'
                        # else: url = match.replace('":"','') + '.jpg'
                        # #print(f'this is cleaned match (URL): {url}')
                        # url_list.append(url)
                    # url_list = list(set(url_list))
                    
                    # finded_urls = len(url_list)
                    # data = {'url_list':url_list, 'total_pics':total_pics , 'finded_urls':finded_urls}
                    # return data
                    
                if '{' in match:
                    url = match.replace('":{"','') + '.jpg'
                else:
                    url = match.replace('":"','') + '.jpg'
                    url_list.append(url)
                    #print(' This url has been findend and appended to list',url) 
        
        
        finded_urls = len(url_list)
        data = {'url_list':url_list, 'total_pics':total_pics , 'finded_pics':finded_urls}
        #print('sent thid data:', data)
        return data

    except Exception as e:
        print('error in get_pictures_url()')
        print(e)
        return 'request error'


proxy_error_count_0 = 0
proxy_error_count_1 = 0
proxy_error_count_2 = 0
proxy_error_count_3 = 0
def proxy_request(url, item):
    ''' url, item'''
    global proxy_error_count_0 
    global proxy_error_count_1 
    global proxy_error_count_2 
    global proxy_error_count_3 
    '''url // downgrade urlib'''
    
    proxy_list = ['193.43.119.41',
                  '193.43.119.3',
                  '5.188.183.221',
                  '5.188.181.86']
    
    # BUG in URLLIB YOU HAVE TO DOWNGRADE URLLIB:  pip install urllib3==1.25.8
    port = '34512'
    user = 'vadymkozak3S7'
    password = 'D9m2IkU'

    session = HTMLSession()
    headers = {'user-agent':"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.159 Safari/537.36"}
    
    #Try with proxies, if error try another and add error to list, if errors list > 100, pop() that proxy
    retries = 0
    n_proxys = len(proxy_list)
    while retries < n_proxys:
        try:
            #proxy = 'http://username:password@ip:port'
            proxy = 'https://'+ user +':'+ password +'@'+ proxy_list[retries] +':'+ port
             
            r = session.get(url, headers=headers, allow_redirects=True, proxies={ 'http':proxy })#proxies={ 'http':proxy,'https':proxy})
            status = r.status_code
            
            if status == 200:
                return r #this breaks the loop
                 
            #All proxies tested without 200 code
            elif status != 200 and retries == n_proxys:
                print(f'ALL proxies tested, any 200 code, status: {status}, item: {item}, URL: {url}') 
                write_bad_result(url=url, item=item)
                return 'request error'
            #still proxy left to retry:
            elif status != 200:
                print(f'error proxy, status: {status}, proxy: {proxy_list[retries]}, URL: {url}') 
                retries += 1
                
                #to maintain only healthy proxies 
                if   retries == 0: proxy_error_count_0 += 1
                elif retries == 1: proxy_error_count_1 += 1
                elif retries == 2: proxy_error_count_2 += 1
                elif retries == 3: proxy_error_count_3 += 1

                if   proxy_error_count_0 > 100: proxy_list.pop(0)
                elif proxy_error_count_1 > 100: proxy_list.pop(1)
                elif proxy_error_count_2 > 100: proxy_list.pop(2)
                elif proxy_error_count_3 > 100: proxy_list.pop(3)
            
        except Exception as e:
            
            print('exception in proxy_request: ',e)
            return 'request error'
    

n_write = 1
def write_processed_item(item,item_link):
    '''item, item_link // write when action performed without errors'''
    global n_write

    wb = load_workbook(filename=output_file)
    ws = wb.active
    
    ws.cell(row=n_write, column=1, value=item)
    ws.cell(row=n_write, column=2, value=item_link)
       
    n_write += 1
    wb.save(output_file)
    # with open('URLs.csv','w') as f:
        # fieldnames = ['Title','URL']
        # writer = csv.DictWriter(f,fieldnames=fieldnames)
        # writer.writeheader()
        
        # n = 0 #used to identify the picture 
        # for item_url in url_list:
            # title = 'query_test' + str(n)
            # n +=1
            # writer.writerow({'Title':title,'URL':item_url})


def send_mega(url, total_pics, item):
    ''' send all files from pics_folder to mega'''
    try:
        #crete a list with all the pics file paths
        file_list = []
        for root, dirs, files in os.walk(pics_folder):
            for file in files:
                file_path = root + file 
                file_list.append(file_path)

        mega = Mega()
        m = mega.login(email, password)
        folder = m.find('V2_IMAGES')
        #upload files
        uploaded_pics = 0
        for picture in file_list:
            retries = 0
            while retries < 4:
                try:
                    #m.upload(file, folder[0])
                    upload = m.upload(picture, folder[0])
                    mega_file_link = m.get_upload_link(upload) 
                    uploaded_pics += 1
                    write_processed_item(picture, mega_file_link)
                    #print('uploaded to mega this file: ', item_name)
                    break
                except Exception as e: #'json.decoder.JSONDecodeError'
                    print(f'Exception in send_mega(): {traceback.print_exc()}')
                    retries += 1
                    if retries == 4: 
                        # retries = 0
                        write_bad_result(url, file)
                        print(f'couldn\'t upload this file to mega {file}')
                        
                        #move the file to a subfloder to avoid delete it and have to download it again later
                        try:
                            print('trying to move pics to errors folder')
                            current_pic_path = pics_folder + '/' + file
                            new_path = pics_folder + '/' + errors_folder
                            shutil.move(current_pic_path, new_path)
                        except Exception as e:
                            print('coulnd\'t move the file to new folder')
                            print(e)

        #print(f'--log_record: {item_name}- {total_pics}---{uploaded_pics} - // total-uploaded')
        #print with formated fields to even space the print
        # print(f'--log_record: {item:30s}{int(total_pics):4.1f}--{int(uploaded_pics):4.2f}')
        print(f'--log_record: {item:30s}{total_pics:4.1f}--{uploaded_pics:4.2f}')
    except Exception as e:
        print('-------------------error while uploading files to mega-------------')
        print(f'this item <{file}>', e)
        traceback.print_exc()
        

def delete():
    '''deletes all the files in pics_folder, to declutter it'''
    try:
        filelist = [ f for f in os.listdir(pics_folder)]#if f.endswith(".bak") ]
        for f in filelist:
            os.remove(os.path.join(pics_folder, f))
    except Exception as e:
        print('error in delete()')
        print(e)
        pass


def run():
    targets_list = extract() #get items and url's from excel file
    for e in targets_list:
        try:
            item = e.get('item')
            url = e.get('url')

            if 'amazon'not in url and 'backmarket' not in url:
                print('not this one!')
                write_bad_result(url, item)
                continue

            r = proxy_request(url, item)

            if r == 'request error':
                logging.debug('request error in proxy_request()')
                continue
            
            data = get_pictures_urls(url, r)

            url_list = data.get('url_list')
            total_pics = data.get('total_pics')
            
            if url_list == 'request error':
                print('request error with this item {} and this url {} continue to next'.format(item,url))
                continue

            n_name = 0 # used to name the downloads, iphone_verde_1, verde_2....
            for pic in url_list:
                if pic == None :
                    print('in run loop, founded None url')
                    continue

                download_picture(pic, item, n_name) #item used to set the name
                n_name += 1
                sleep(3)
            send_mega(url, total_pics, item)#only URL needed, the rest used for logging

            delete()

        except Exception as e:
            print('error in run loop:', e)
            #traceback.print_exc()
            continue


if __name__ == "__main__":

    run()