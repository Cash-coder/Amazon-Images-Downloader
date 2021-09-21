from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

''' this code takes a excel with values marked with an X, 
and take those values to a new excel only if the value is unique, note repeated'''

input_file = 'phones_matches.xlsx'
output_file = 'phones_set.xlsx'


def extract():
    ''' reads input file and create a dict with item and url,only unique URL's'''
    wb = load_workbook(filename = input_file)
    ws = wb.active

    target_list = []
    burned_list = []
    for row in ws.iter_rows(values_only=True):
        item = row[0]
        url = row[4]
        match = row[3]
        if match != None:
            match = match.lower()

        if match == 'x' and item not in burned_list :#and 'amazon' in url:
            print(item)
            entry = {'item':item, 'url':url}
            target_list.append(entry)
            burned_list.append(item)

    return target_list


n = 1
def write_file(url_list):
    ''' writes the list into a new excel file, this time only with unique values'''
    global n

    wb = load_workbook(filename=output_file)
    ws = wb.active

    for url in url_list:#ws.iter_rows(values_only=True):
        
        item = url.get('item')
        url = url.get('url')
        
        ws.cell(row=n, column=1, value=item)
        ws.cell(row=n, column=2, value=url)
        n += 1
        wb.save(output_file)


def run():

    set_list = extract()

    write_file(set_list)


if __name__ == "__main__":
    run()