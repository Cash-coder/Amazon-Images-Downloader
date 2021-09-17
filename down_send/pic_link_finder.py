#https://pypi.org/project/mega.py/
from mega import Mega
import login_file # .py file with MEGA user and password
from requests_html import HTMLSession
import requests
import re
from time import sleep
from requests_html import HTMLSession
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
import os

input_file = 'amazon_list.xlsx'
output_file = 'links_founded.xlsx'
errors_file = 'errors.xlsx'

pics_folder =  r'C:/Users/HP EliteBook/OneDrive/A_Miscalaneus/Escritorio/Code/git_folder/images_downloader/down_send/pics_folder/'

#returns a list of dicts with {item , url}
def extract():
    wb = load_workbook(filename = input_file)
    ws = wb.active
    target_list = []
    #item_attribute_list = []
    for row in ws.iter_rows(values_only=True):
        item = row[0]
        url = row[1]

        entry = {'item':item, 'url':url}
        target_list.append(entry)
        # if 'GB' in item:
        #     chain = item.split(' ')
        #     item_p = chain.remove(chain[-1])
        #     item_p = chain.remove(chain[-2])
        #     print(item_p)
        #     #myList.remove(myList[len(myList)-1])
    return target_list

# def make_query(item):
#     query = item.replace(' ', '+')
#     return query

n_errors = 1
def write_bad_result(item,url):
    global n_errors

    wb = load_workbook(filename = errors_file)
    ws = wb.active

    ws.cell(row=n,column=1,value=item)
    ws.cell(row=n,column=2,value=url)

    n_errors += 1


n = 0
def download_picture(url,item): # Download To local machine
    global n
    global pics_folder
    try:
        pic = requests.get(url)

        if pic.status_code == 200:
            file_name = item + ' ' + str(n) +'.jpg'
            file_name = file_name.replace(' ','_')
    
            try:
                
                complete_path = pics_folder + file_name
                
                with open(complete_path, 'wb') as f:
                    f.write(pic.content)
                n += 1
            except Exception as e:
                print(e)
                print('failed to save this pic',item, url)
                write_bad_result(item,url)
        else:
            write_bad_result(item,url)
            
    except Exception as e:
        print(e)
        write_bad_result(item,url)
        save_html_response(pic,item)


        
def save_html_response(r,item):
    item = item.replace(' ', '_')
    with open(item + 'response.html','w',encoding='utf-8') as f:
        f.write(r.text)


def get_pictures_urls(s):
    url_list = []
    
    # hiRes == High Resolution 
    # main  ==  ?
    # large ==  ?
    matches = re.findall(r'hiRes(.*?).jpg',s)
    for match in matches:
        if '{' in match:
            url = match.replace('":{"','') + '.jpg'
        else:
            url = match.replace('":"','') + '.jpg'
        url_list.append(url)  
        print('url findend and appended to list',url) 
    return url_list

def raw_request(url):

    headers = {'user-agent':"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.159 Safari/537.36"}
    session = HTMLSession()

    r = session.get(url,headers=headers,allow_redirects=True)
    status = r.status_code
    #To visualize response in the browser
    #save_html_response(r)
    if status == 200:
        tag = r.html.find('script', containing='P.when(\'A\').register("ImageBlockATF", function(A){')
        s = tag[0].text
    else:
        print("Response status:{} for the url: {}".format(status,url))

    return s

n = 1
def write_file(url_list,item):
    global n

    wb = load_workbook(filename=output_file)
    ws = wb.active

    for url in url_list:#ws.iter_rows(values_only=True):
        ws.cell(row=n, column=1, value=item)
        ws.cell(row=n, column=2, value=url)
        n += 1
        wb.save(output_file)
    # with open('URLs.csv','w') as f:
        # fieldnames = ['Title','URL']
        # writer = csv.DictWriter(f,fieldnames=fieldnames)
        # writer.writeheader()
        
        # n = 0 #used to identify the picture 
        # for item_url in url_list:
            # title = 'query_test' + str(n)
            # n +=1
            # writer.writerow({'Title':title,'URL':item_url})

def send_mega():
    #crete a list with all the file paths
    file_list = []
    for root, dirs, files in os.walk(pics_folder):
        for file in files:
            file_path = root + file 
            file_list.append(file_path)

    #init mega
    email = login_file.email
    password = login_file.password

    mega = Mega()
    m = mega.login(email, password)
    folder = m.find('IMAGES')
    #upload files
    for file in file_list:
        m.upload(file, folder[0])

def delete():
    filelist = [ f for f in os.listdir(pics_folder)]#if f.endswith(".bak") ]
    for f in filelist:
        os.remove(os.path.join(pics_folder, f))

def run():
    pass


#url = 'https://www.amazon.es/s?k='+ query +'&__mk_es_ES=%C3%85M%C3%85%C5%BD%C3%95%C3%91&ref=nb_sb_noss'
#url = 'https://www.amazon.es/Apple-iPhone-64GB-Plata-Reacondicionado/dp/B082DKM8TG/ref=sr_1_3?__mk_es_ES=%C3%85M%C3%85%C5%BD%C3%95%C3%91&crid=3L65HJLXS1PD9&dchild=1&keywords=iphine11&qid=1630505792&s=electronics&sprefix=iphi%2Celectronics%2C185&sr=1-3'
    #print(s)
# query = 'iphone 12 pro grafito'
# query = query.replace(' ','+')


targets_list = extract()



for e in targets_list:

    item = e.get('item')
    url = e.get('url')

    #query = make_query(item)
    raw_data = raw_request(url)
    url_list = get_pictures_urls(raw_data)
    #for url in url_list:
        #download_picture(url)
    #write_file(url_list,item)
    for pic in url_list:
        download_picture(pic, item) #item used to set the name
        sleep(3)
    
    send_mega()

    delete()


#when all processed, download the links, 
# title qith the query n+1
# send to mega
#Gustavo review 50 by 50 HOW MARK THE MISSING ONES ?
# upload the images files or the image url's of mega to prod_db
# prod color pictures



############### Maybe useful code later: ################
#save_html_response(r)

#soup = bs4(r.text, 'lxml')
#soup = bs4(r.content, 'lxml')
#soup = bs4(tag[0].text,'lxml')

#line of code to identify element with the pictures
'''
P.when('A').register("ImageBlockATF", function(A){
'''