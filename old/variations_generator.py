import itertools
from openpyxl import load_workbook
import json

from openpyxl.workbook.workbook import Workbook 

#wb = Workbook()
wb = load_workbook('phones.xlsx')
ws = wb.active
#print("loaded wb----------")

prod_states = 'nuevo','seminuevo'
variations_list = []

for row in ws.iter_rows(values_only=True):
    #print(row,'---------------------')
    if row[0] == 'PHONE':
        #identify the values
        brand = row[1]
        model = str(row[1])
        capacity = row[3]
        colors = row[3]
        #print(brand,model,capacity,colors)
        #extract lists like 64GB,128GB,256GB
        #some phones only have one capacity or color
        try:
            colors = colors.split(' ')
            #print(colors)
        except Exception as e:
            print(e)
            pass
        
        if colors and colors != ' ':
            for color in colors:
                variation = model + ' ' + color
                # try:
                #     variation = variation.replace('.0','').replace(',','')
                # except:
                #     pass

                #item = brand + ' ' + model
                variations_list.append({'variation':variation,'item':model,'color':color})
                print(variation)
        

        # try:
        #     capacity = capacity.split(' ')
        #     #print(capacity)
        #     #If GB not present in capacity, add it
        #     if 'GB' not in str(capacity):
        #         capacity_list=[]
        #         for item in capacity:
        #             item = item + 'GB'
        #             capacity_list.append(item)
        #             #print(item)
        #         #capacity.clear()
        #         for item in capacity_list:
        #             capacity.append(item)
        #         #print(capacity_list)
        # except Exception as e:
        #     print(e)
        #     pass

        # print(model,colors)
        # try:
        #     combinations = list(itertools.product(colors,str(model)))#,capacity_list))
        #     #print(combinations)

        #     for combination in combinations:
        #         #print(combination)
        #         #combination = list(combination)
        #         #one variation for nuevo, other for seminuevo
        #         variation1 = brand + ' ' + combination[1] + ' ' + combination[0]
        #         #variation2 = brand + ' ' + model + ' ' + combination[1] + ' ' + combination[0] + ' ' + prod_states[1]
        #         #variation1 = brand + ' ' + model + ' ' + combination[1] + ' ' + combination[0] + ' ' + prod_states[0]
        #         # print(variation1)
        #         # print(variation2)
        #         variations_list.append(variation1)
        #         #variations_list.append(variation2)
        # except Exception as e:
        #     print(e)
        #     pass
    #avoid blank rows
    elif row == None:
        continue

#writing to a new excel file
wb2 = Workbook()
ws2 = wb2.active

ws = wb.active
n = 1
for e in variations_list:
    
    variation = e.get('variation')
    item = e.get('item')
    color = e.get('color')

    ws2.cell(row=n,column=1).value = item
    ws2.cell(row=n,column=2).value = color
    ws2.cell(row=n,column=3).value = variation
    n +=1
# for row in ws2.iter_rows():
#     value = variations_list.pop()   
#     print(value)
#     row = 
wb2.save('tablets_color_variations.xlsx')


# def get_phone_variations(brand,model,colors, capacity):
    
#     combinations = list(itertools.product(colors,capacity))
#     #print(combinations)
#     variations_list = []

#     for combination in combinations:
#         combination = list(combination)
#         #one variation for nuevo, other for seminuevo
#         variation2 = brand + ' ' + model + ' ' + ' ' + combination[1] + ' ' + combination[0] + ' ' + prod_states[1]
#         variation1 = brand + ' ' + model + ' ' + ' ' + combination[1] + ' ' + combination[0] + ' ' + prod_states[0]
#         variations_list.append(variation1)
#         variations_list.append(variation2)
        

#     return variations_list

# prod_states = ['nuevo', 'seminuevo'] # iphone nuevo, iphone seminuevo

# target_products_seed = [ #This is the short version,all the variations are in target_db.json

#     {'product':'phone', 'brand':'iphone', 'model':'12', 'colors': ('verde', 'azul'), 'capacity':('64GB','128GB', '256GB')},
#     #{'product':'phone','brand':'iphone', 'model':'11', 'colors': ('amarillo', 'rojo'), 'capacity':('64GB','128GB', '256GB')},

#     ]

# dict_variations = {0:{'name': 'test_name', 'quantity':0}}
# list_variations = []

# with open('target_db_2.json', 'w') as jsonfile:
#     json.dump(dict_variations,jsonfile,indent=4)

# for prod in target_products_seed:
    
#     if prod['product'] == 'phone':
#         brand = prod['brand']
#         model = prod['model']
#         colors = list(prod['colors'])
#         capacity = list(prod['capacity'])

#         variations = get_phone_variations(brand,model,colors,capacity)        
        
#         #dict_variations = { i : 0 for i in variations }

#         print(variations)

#         n=1
#         for variation in variations:
#             adict = {n:{'name':variation,'quantity':0}}
#             n +=1
#             with open('target_db_2.json', 'r+') as jsonfile:
#                 current_data = json.load(open('target_db_2.json'))

#                 current_data.update(adict)
#                 json.dump(adict,jsonfile)



#         # #############################

#         # for variation in variations:            
#         #     # dict_variations['name'] = variation
#         #     # dict_variations['quantity'] = 0

#         #     dict_variations.update({'name':variation,'quantity':0})
#         #     print('dict_variations',dict_variations)
            
#         #     list_variations.append(dict_variations)
#         #     #curret_data = json.loads('target_db.json')

            
#         #     # with open('target_db_2.json', 'w') as jsonfile:
                
#         #     #     file_data = json.load(jsonfile)
#         #     #     file_data.append(dict_variations)

#         #     #     json.dump(file_data,jsonfile,indent=4)

#         # #print(list_variations)
#         # with open('target_db_2.json', 'r+') as jsonfile:
#         #     for item in list_variations:
#         #         name = item['name']

#         #         dict_variation= {'name':name,'quantity':0}
#         #         json.dump(dict_variation,jsonfile)
            


   